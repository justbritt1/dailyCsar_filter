from flask import Flask, request, render_template, send_file, session, redirect, url_for
import pandas as pd
import json
import io
import os
import openpyxl
import time

app = Flask(__name__)
app.secret_key = 'itsasecretkey'  # Change this to a random secret key

# Columns to compare between uploads
columns_to_compare = ["Sec Faculty Info", "Sec All Faculty Last Names", "Total FTE", "FTE Count"]

# Directory for temporary files
TEMP_DIR = os.path.join(os.path.dirname(__file__), 'temp_data')
if not os.path.exists(TEMP_DIR):
    os.makedirs(TEMP_DIR)


def get_temp_filepath(filename):
    """Generate a safe temporary file path"""
    return os.path.join(TEMP_DIR, filename)


def cleanup_temp_files():
    """Clean up old temporary files"""
    temp_files = [
        'temp_new_data.csv',
        'temp_master.xlsx',
        'preview.html',
        'changes.json',
        'updated_master.csv',
        'updated_master.xlsx'
    ]
    for filename in temp_files:
        filepath = get_temp_filepath(filename)
        if os.path.exists(filepath):
            try:
                os.remove(filepath)
            except: 
                pass


@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        file = request.files['file']
        if file:
            filename = file.filename. lower()
            
            # Read the new file (CSV or Excel)
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                df = pd.read_excel(file)
            else:
                df = pd.read_csv(file)
            
            df.columns = df.columns.str.strip()

            # Save the new data temporarily to file
            temp_csv_path = get_temp_filepath("temp_new_data.csv")
            df.to_csv(temp_csv_path, index=False)
            
            # Save preview HTML to file instead of session
            preview_html = df.head(10).to_html(classes='data')
            preview_path = get_temp_filepath("preview.html")
            with open(preview_path, 'w', encoding='utf-8') as f:
                f.write(preview_html)
            
            # Store only a flag in session (not the data!)
            session['has_new_data'] = True
            
            return redirect(url_for('select_master'))

    return render_template('index.html')


@app.route('/select-master', methods=['GET', 'POST'])
def select_master():
    if not session.get('has_new_data'):
        return redirect(url_for('upload_file'))
    
    start_time = time.time()
    
    if request.method == 'POST':
        master_file = request.files['master_file']
        if master_file:
            master_filename = master_file.filename.lower()
            
            # Load the master file (CSV or Excel)
            if master_filename.endswith('.xlsx') or master_filename.endswith('.xls'):
                is_excel = True
                excel_file_path = get_temp_filepath("temp_master.xlsx")
                master_file.save(excel_file_path)
                master_df = pd.read_excel(excel_file_path)
            else:
                is_excel = False
                excel_file_path = None
                master_df = pd.read_csv(master_file)
            
            master_df.columns = master_df.columns.str.strip()
            
            # Load the new data from file
            temp_csv_path = get_temp_filepath("temp_new_data.csv")
            new_df = pd.read_csv(temp_csv_path)
            new_df.columns = new_df.columns.str.strip()
            
            # Check for Term column
            term_column = None
            if 'Term' in new_df.columns and 'Term' in master_df.columns:
                term_column = 'Term'
            
            # Check if files have a common key column (try common column names)
            key_column = None
            possible_keys = ['Section Name', 'Sec Name', 'ID', 'Section', 'Name']
            
            for key in possible_keys:
                if key in new_df.columns and key in master_df.columns:
                    key_column = key
                    break
            
            if not key_column:
                # If no common key found, use first column
                if len(new_df.columns) > 0 and len(master_df.columns) > 0:
                    if new_df.columns[0] == master_df.columns[0]: 
                        key_column = new_df.columns[0]
            
            if not key_column:
                return "Error: Could not find a matching column between the files to compare."
            
            print(f"Using key column: {key_column}")
            if term_column:
                print(f"Using term column: {term_column}")
            
            # Clean the key columns in both dataframes to improve matching
            master_df[key_column] = master_df[key_column].astype(str).str.strip()
            new_df[key_column] = new_df[key_column].astype(str).str.strip()
            
            if term_column:
                master_df[term_column] = master_df[term_column].astype(str).str.strip()
                new_df[term_column] = new_df[term_column].astype(str).str.strip()
            
            # Filter new_df to only rows with non-empty key column (primary rows only)
            new_df_primary = new_df[new_df[key_column].notna() & (new_df[key_column] != '') & (new_df[key_column] != 'nan')].copy()
            
            print(f"Total rows in new file: {len(new_df)}")
            print(f"Primary rows (with Section Name): {len(new_df_primary)}")
            
            # Detect changes
            changes_list = []
            updated_master = master_df.copy()

            for idx, new_row in new_df_primary.iterrows():
                sec_name = str(new_row[key_column]).strip()
                
                # Use composite key (Term + Section Name) if Term column exists
                if term_column:
                    term_val = str(new_row[term_column]).strip()
                    master_idx = master_df[
                        (master_df[key_column] == sec_name) & 
                        (master_df[term_column] == term_val)
                    ].index
                else:
                    master_idx = master_df[master_df[key_column] == sec_name].index
                
                if len(master_idx) > 0:
                    master_idx = master_idx[0]
                    # Check for changes in specified columns
                    has_changes = False
                    change_details = {}
                    
                    for col in columns_to_compare:
                        if col in new_df.columns and col in master_df.columns:
                            old_val = master_df.loc[master_idx, col]
                            new_val = new_row[col]
                            
                            # Better comparison handling
                            old_str = str(old_val).strip() if pd.notna(old_val) else ''
                            new_str = str(new_val).strip() if pd.notna(new_val) else ''
                            
                            if old_str != new_str:
                                has_changes = True
                                change_details[col] = {
                                    'old': old_str,
                                    'new': new_str
                                }
                    
                    # Update ALL columns from new data to the corresponding row in master
                    for col in new_df.columns:
                        if col in updated_master.columns:
                            updated_master.loc[master_idx, col] = new_row[col]
                    
                    if has_changes:
                        identifier = f"{term_val} - {sec_name}" if term_column else sec_name
                        changes_list.append({
                            'identifier': identifier,
                            'change_type': 'modified',
                            'changes': change_details
                        })
                else:
                    # New section not in master - add it (truly new section)
                    identifier = f"{term_val} - {sec_name}" if term_column else sec_name
                    print(f"New section found: {identifier}")
                    updated_master = pd.concat([updated_master, new_row.to_frame().T], ignore_index=True)
                    changes_list.append({
                        'identifier': identifier,
                        'change_type': 'new',
                        'changes': {}
                    })
            
            # Save updated master
            if is_excel:
                # Load workbook to preserve formatting
                from openpyxl import load_workbook
                wb = load_workbook(excel_file_path)
                ws = wb.active
                
                # Create a mapping of column names to Excel column indices in master
                col_name_to_idx = {}
                header_row = [cell.value for cell in ws[1]]
                for col_idx, col_name in enumerate(header_row):
                    if col_name: 
                        col_name_to_idx[str(col_name).strip()] = col_idx + 1  # Excel is 1-indexed
                
                # Update cells with new data - ONLY PRIMARY ROWS
                for idx, new_row in new_df_primary.iterrows():
                    sec_name = str(new_row[key_column]).strip()
                    
                    # Use composite key (Term + Section Name) if Term column exists
                    if term_column:
                        term_val = str(new_row[term_column]).strip()
                        master_idx = master_df[
                            (master_df[key_column] == sec_name) & 
                            (master_df[term_column] == term_val)
                        ].index
                    else:
                        master_idx = master_df[master_df[key_column] == sec_name].index
                    
                    if len(master_idx) > 0:
                        row_num = master_idx[0] + 2  # +2 because Excel is 1-indexed and has header
                        
                        # Update each column that exists in both files
                        for col_name in new_df.columns:
                            col_name_str = str(col_name).strip()
                            if col_name_str in col_name_to_idx:
                                excel_col = col_name_to_idx[col_name_str]
                                cell_value = new_row[col_name]
                                # Handle NaN values
                                if pd.isna(cell_value):
                                    cell_value = None
                                ws.cell(row=row_num, column=excel_col).value = cell_value
                
                output_path = get_temp_filepath("updated_master.xlsx")
                wb.save(output_path)
                output_ext = '.xlsx'
            else:
                output_path = get_temp_filepath("updated_master.csv")
                updated_master.to_csv(output_path, index=False)
                output_ext = '.csv'
            
            # Right before saving changes to JSON:
            processing_time = round(time.time() - start_time, 2)

            # Save changes to JSON file instead of session
            changes_path = get_temp_filepath("changes.json")
            with open(changes_path, 'w', encoding='utf-8') as f:
                json.dump(changes_list, f)
            
            # Store only metadata in session
            session['has_new_data'] = False
            session['output_ext'] = output_ext
            session['num_changes'] = len(changes_list)
            session['processing_time'] = processing_time
            
            return redirect(url_for('results'))
    
    # Load preview from file
    preview_path = get_temp_filepath("preview.html")
    preview_html = ''
    if os.path.exists(preview_path):
        with open(preview_path, 'r', encoding='utf-8') as f:
            preview_html = f.read()
    
    return render_template('select_master.html', preview=preview_html)



@app.route('/results')
def results():
    """Display results page"""
    num_changes = session.get('num_changes', 0)
    processing_time = session.get('processing_time', 0)
    
    # Load changes from file
    changes_path = get_temp_filepath("changes.json")
    changes_list = []
    if os.path.exists(changes_path):
        with open(changes_path, 'r', encoding='utf-8') as f:
            changes_list = json.load(f)
    
    # Load alerts from file if they exist
    alerts_path = get_temp_filepath("alerts.json")
    alerts_list = []
    if os.path.exists(alerts_path):
        with open(alerts_path, 'r', encoding='utf-8') as f:
            alerts_list = json.load(f)
    
    # Calculate num_rows_processed from the saved new data file
    temp_csv_path = get_temp_filepath("temp_new_data.csv")
    num_rows_processed = 0
    if os.path.exists(temp_csv_path):
        try:
            df = pd.read_csv(temp_csv_path)
            num_rows_processed = len(df)
        except Exception as e:
            print(f"Error reading temp data: {e}")
            num_rows_processed = 0
    
    # Calculate statistics from changes_list
    num_new = 0
    num_modified = 0
    num_deleted = 0
    
    for change in changes_list:
        change_type = change.get('change_type', '').lower()
        if change_type == 'new':
            num_new += 1
        elif change_type == 'modified':
            num_modified += 1
        elif change_type == 'deleted':
            num_deleted += 1
    
    return render_template('results.html', 
                         num_changes=num_changes,
                         num_new=num_new,
                         num_modified=num_modified,
                         num_deleted=num_deleted,
                         processing_time=processing_time,
                         changes_list=changes_list,
                         alerts_list=alerts_list,
                         num_rows_processed=num_rows_processed)


@app.route('/download')
def download():
    output_ext = session.get('output_ext', '.csv')
    output_filename = 'updated_master' + output_ext
    output_path = get_temp_filepath(output_filename)
    
    if os.path.exists(output_path):
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' if output_ext == '.xlsx' else 'text/csv'
        return send_file(output_path, mimetype=mimetype, as_attachment=True, download_name=output_filename)
    return "No updated file found."


@app.route('/cleanup')
def cleanup():
    """Manual cleanup endpoint"""
    cleanup_temp_files()
    return redirect(url_for('upload_file'))


if __name__ == '__main__':
    app.run(debug=True)