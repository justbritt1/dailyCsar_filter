from flask import Flask, request, render_template, send_file, session, redirect, url_for
import pandas as pd
import io
import os
import openpyxl

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'  # Change this to a random secret key

# Columns to compare between uploads
columns_to_compare = ["Sec Faculty Info", "Sec All Faculty Last Names", "Total FTE", "FTE Count"]

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        file = request.files['file']
        if file:
            filename = file.filename.lower()
            
            # Read the new file (CSV or Excel)
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                df = pd.read_excel(file)
            else:
                df = pd.read_csv(file)
            
            df.columns = df.columns.str.strip()

            # Save the new data temporarily
            df.to_csv("temp_new_data.csv", index=False)
            
            # Store in session that we have new data
            session['has_new_data'] = True
            session['new_data_preview'] = df.head(10).to_html(classes='data')
            
            return redirect(url_for('select_master'))

    return render_template('index.html')

@app.route('/select-master', methods=['GET', 'POST'])
def select_master():
    if not session.get('has_new_data'):
        return redirect(url_for('upload_file'))
    
    if request.method == 'POST':
        master_file = request.files['master_file']
        if master_file:
            master_filename = master_file.filename.lower()
            
            # Load the master file (CSV or Excel)
            if master_filename.endswith('.xlsx') or master_filename.endswith('.xls'):
                is_excel = True
                excel_file_path = os.path.join(os.path.dirname(__file__), "temp_master.xlsx")
                master_file.save(excel_file_path)
                master_df = pd.read_excel(excel_file_path)
            else:
                is_excel = False
                excel_file_path = None
                master_df = pd.read_csv(master_file)
            
            master_df.columns = master_df.columns.str.strip()
            
            # Load the new data
            new_df = pd.read_csv("temp_new_data.csv")
            new_df.columns = new_df.columns.str.strip()
            
            # Check if files have a common key column (try common column names)
            key_column = None
            possible_keys = ['Sec Name', 'Section Name', 'ID', 'Section', 'Name']
            
            for key in possible_keys:
                if key in new_df.columns and key in master_df.columns:
                    key_column = key
                    break
            
            if not key_column:
                # If no common key found, use first column
                if len(new_df.columns) > 0 and len(master_df.columns) > 0:
                    if new_df.columns[0] == master_df.columns[0]:
                        key_column = new_df.columns[0]
            
            if not key_column:
                return "Error: Could not find a matching column between the files to compare."
            
            # Detect changes
            changes_list = []
            updated_master = master_df.copy()
            
            for idx, new_row in new_df.iterrows():
                sec_name = new_row[key_column]
                master_idx = master_df[master_df[key_column] == sec_name].index
                
                if len(master_idx) > 0:
                    master_idx = master_idx[0]
                    # Check for changes in specified columns
                    has_changes = False
                    change_info = {key_column: sec_name}
                    
                    for col in columns_to_compare:
                        if col in new_df.columns and col in master_df.columns:
                            old_val = master_df.loc[master_idx, col]
                            new_val = new_row[col]
                            if old_val != new_val:
                                has_changes = True
                                change_info[col + ' (Old)'] = old_val
                                change_info[col + ' (New)'] = new_val
                    
                    # Update ALL columns from new data to the corresponding row in master
                    for col in new_df.columns:
                        if col in updated_master.columns:
                            updated_master.loc[master_idx, col] = new_row[col]
                    
                    if has_changes:
                        changes_list.append(change_info)
                else:
                    # New row not in master - add it
                    updated_master = pd.concat([updated_master, new_row.to_frame().T], ignore_index=True)
                    change_info = {key_column: sec_name, 'Status': 'NEW ROW ADDED'}
                    changes_list.append(change_info)
            
            # Save updated master
            if is_excel:
                # Load workbook to preserve formatting
                from openpyxl import load_workbook
                wb = load_workbook(excel_file_path)
                ws = wb.active
                
                # Update cells with new data
                for idx, new_row in new_df.iterrows():
                    sec_name = new_row[key_column]
                    master_idx = master_df[master_df[key_column] == sec_name].index
                    
                    if len(master_idx) > 0:
                        row_num = master_idx[0] + 2  # +2 because Excel is 1-indexed and has header
                        
                        # Update each column
                        for col_idx, col_name in enumerate(new_df.columns):
                            if col_name in master_df.columns:
                                excel_col = col_idx + 1
                                ws.cell(row=row_num, column=excel_col, value=new_row[col_name])
                
                output_path = os.path.join(os.path.dirname(__file__), "updated_master.xlsx")
                wb.save(output_path)
                output_ext = '.xlsx'
            else:
                output_path = os.path.join(os.path.dirname(__file__), "updated_master.csv")
                updated_master.to_csv(output_path, index=False)
                output_ext = '.csv'
            
            # Prepare results
            changes_df = pd.DataFrame(changes_list) if changes_list else pd.DataFrame()
            
            session['has_new_data'] = False
            session['output_file'] = output_path
            session['output_ext'] = output_ext
            
            return render_template('results.html', 
                                 num_changes=len(changes_list))
    
    return render_template('select_master.html', 
                         preview=session.get('new_data_preview'))
@app.route('/download')
def download():
    output_path = session.get('output_file')
    output_ext = session.get('output_ext', '.csv')
    
    if output_path and os.path.exists(output_path):
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' if output_ext == '.xlsx' else 'text/csv'
        download_name = 'updated_master' + output_ext
        return send_file(output_path, mimetype=mimetype, as_attachment=True, download_name=download_name)
    return "No updated file found."

if __name__ == '__main__':
    app.run(debug=True)